import csv
from dataclasses import dataclass
import functools
import json
import subprocess
import sys
from textwrap import dedent
import typing
import zipfile
from io import SEEK_SET
from pathlib import Path
from typing import Annotated, Any, Dict, List, Optional, TextIO, Tuple

from flytekit.core.annotation import FlyteAnnotation
from latch import medium_task, workflow
from latch.resources.launch_plan import LaunchPlan
from latch.types import (
    LatchDir,
    LatchFile,
    LatchMetadata,
    LatchAuthor,
    LatchOutputDir,
    LatchParameter,
    LatchRule,
    Section,
    Fork,
    ForkBranch,
    Text,
    Params,
)
from latch.types.metadata import FlowBase
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException

from wf.report_gen import generate_report
from wf.util import error, message, warn, warning

sys.stdout.reconfigure(line_buffering=True)

# Strips byte order marks sometimes present at beginning of tabular files
# exported from excel
functools.partial(open, encoding="utf-8-sig")

csv.field_size_limit(sys.maxsize)


def csv_tsv_reader(f: TextIO, use_dict_reader: bool = False):
    sniff = csv.Sniffer()
    dialect = sniff.sniff(f.readline())
    f.seek(0, SEEK_SET)

    if use_dict_reader:
        return csv.DictReader(f, dialect=dialect)
    else:
        return csv.reader(f, dialect=dialect)


def pull_gene_from_header(csv: Path) -> Optional[str]:
    with open(csv) as f:
        r = csv_tsv_reader(f)
        return next(r)[0]


@medium_task
def deseq2(
    report_name: str,
    count_table_source: str = "single",
    raw_count_table: Optional[LatchFile] = None,
    raw_count_tables: List[LatchFile] = [],
    count_table_gene_id_column: Optional[str] = None,
    output_location_type: str = "default",
    output_location: Optional[LatchDir] = None,
    conditions_source: str = "manual",
    manual_conditions: List[List[str]] = [],
    conditions_table: Optional[LatchFile] = None,
    design_matrix_sample_id_column: Optional[str] = None,
    design_formula: List[List[str]] = [["condition", "explanatory"]],
    number_of_genes_to_plot: int = 30,
) -> LatchDir:

    # Hack until proper string conditionals exist on bulk
    if conditions_source == "none":
        return LatchDir("/root/wf")

    if count_table_gene_id_column is None:
        count_table_gene_id_column = "gene_id"

    if design_matrix_sample_id_column is None:
        design_matrix_sample_id_column = "sample_id"

    if conditions_source == "manual":
        design_formula = [["condition", "explanatory"]]

    if len(design_formula) == 0:
        raise ValueError("design formula is empty")

    design_formula_explanatory = []
    design_formula_confounding = []
    design_formula_cluster = []
    for x in design_formula:
        if x[1] == "explanatory":
            design_formula_explanatory.append(x[0])
        if x[1] == "confounding":
            design_formula_confounding.append(x[0])
        if x[1] == "cluster":
            design_formula_cluster.append(x[0])

    if count_table_source == "single":
        if raw_count_table is None:
            raise ValueError("Expected the single count table source to be set")
        count_table_remote = raw_count_table.remote_source
        raw_count_table_p = Path(raw_count_table)
        count_table_gene_id_column = pull_gene_from_header(raw_count_table_p)
    else:
        raw_count_table_p = Path("combined_counts.csv")
        with raw_count_table_p.open("w") as f:
            w = csv.writer(f)

            input_fds = [Path(x).open("r") for x in raw_count_tables]
            try:
                readers = [csv_tsv_reader(f) for f in input_fds]

                for line_idx, line in enumerate(readers[0]):
                    if line_idx == 0:
                        count_table_gene_id_column = line[0]

                    for r in readers[1:]:
                        other_line = next(r)

                        if other_line[0] != line[0]:
                            raise ValueError(
                                "Counts tables being combined are not sorted in the"
                                f" same order: {other_line[0]} != {line[0]} on line"
                                f" {line_idx}"
                            )

                        line.extend(other_line[1:])

                    w.writerow(line)
            finally:
                for f in input_fds:
                    f.close()

        count_table_remote = "combined"

    if output_location_type == "custom" and output_location is None:
        error(
            {
                "title": "Invariant violation",
                "body": "Expected a custom output location but it is null",
            }
        )
        raise RuntimeError("custom output location requested but not specified")

    if conditions_source == "table" and conditions_table is None:
        error(
            {
                "title": "Invariant violation",
                "body": "Expected a design matrix path but it is null",
            }
        )
        raise RuntimeError(
            "design matrix file input requested but no location specified"
        )

    print(
        ">>> Parameters",
        f"Count table: '{count_table_remote}'",
        f"Report name: '{report_name}'",
        f"Number of Genes: '{str(number_of_genes_to_plot)}'",
        sep="\n",
    )

    output_loc = f"latch:///DESeq2 Results/{report_name.replace('/', '_')}"
    if output_location_type == "custom":
        assert output_location is not None
        output_loc = output_location.remote_path

    print(f"Output location: '{output_loc}' [{output_location_type}]")

    conditions_table_p: Optional[Path] = None
    if conditions_source == "table":
        assert conditions_table is not None
        conditions_table_p = Path(conditions_table)
        print(f"Design matrix file: '{conditions_table.remote_source}'")
    else:
        design_matrix_sample_id_column = "sample_id"
        if len(manual_conditions) == 0:
            error(
                {
                    "title": "Design matrix is empty",
                    "body": "No data provided in the form",
                }
            )
            raise RuntimeError("Design matrix is empty")

        conditions_table_p = Path("conditions.csv")
        with conditions_table_p.open("w") as f:
            w = csv.DictWriter(
                f, fieldnames=[design_matrix_sample_id_column, "condition"]
            )
            w.writeheader()
            w.writerows(
                {
                    design_matrix_sample_id_column: cond[0],
                    "condition": cond[1],
                }
                for cond in manual_conditions
            )

    print()
    excel_okay = False
    genes = set()
    try:
        workbook = load_workbook(str(raw_count_table_p))
        sheet = workbook.worksheets[0]

        headers = None
        gene_id_column_idx = None
        for row_ in sheet.iter_rows():
            row = typing.cast(Tuple[Cell], row_)
            if headers is None:
                headers = row

                for idx, h in enumerate(headers):
                    if h.value != count_table_gene_id_column:
                        continue
                    gene_id_column_idx = idx
                    break

                if gene_id_column_idx is None:
                    error(
                        {
                            "title": "Invalid gene ID column selected",
                            "body": [
                                f"Gene ID column '{count_table_gene_id_column}'"
                                " could not be found",
                                {
                                    "section": "Available Columns:",
                                    "body": {"list": [str(h.value) for h in headers]},
                                },
                            ],
                        }
                    )
                    raise RuntimeError("Invalid sample ID column")

                continue

            assert gene_id_column_idx is not None

            genes.add(row[gene_id_column_idx].value)

        excel_okay = True
    except InvalidFileException:
        pass
    except zipfile.BadZipFile:
        pass

    if not excel_okay:
        with raw_count_table_p.open("r") as f:
            r = csv_tsv_reader(f, use_dict_reader=True)
            for row in r:
                items = row.items()
                genes.add(row[count_table_gene_id_column])

    print()
    print(
        "Design matrix:"
        f" [{'table file' if conditions_source == 'table' else 'manual input'}]"
    )
    excel_okay = False
    try:
        workbook = load_workbook(str(conditions_table_p))
        sheet = workbook.worksheets[0]

        headers = None
        sample_id_column_idx = None
        for row_ in sheet.iter_rows():
            row = typing.cast(Tuple[Cell], row_)
            if headers is None:
                headers = row

                for idx, h in enumerate(headers):
                    if h.value != design_matrix_sample_id_column:
                        continue
                    sample_id_column_idx = idx
                    break

                if sample_id_column_idx is None:
                    error(
                        {
                            "title": "Invalid sample ID column selected",
                            "body": [
                                f"Sample ID column '{design_matrix_sample_id_column}'"
                                " could not be found",
                                {
                                    "section": "Available Columns:",
                                    "body": {"list": [str(h.value) for h in headers]},
                                },
                            ],
                        }
                    )
                    raise RuntimeError("Invalid sample ID column")

                continue

            assert sample_id_column_idx is not None

            print(
                f"{row[sample_id_column_idx].value}: "
                f"{', '.join(str(x.value) for idx, x in enumerate(row) if idx != sample_id_column_idx)}"
            )

        excel_okay = True
    except InvalidFileException:
        pass
    except zipfile.BadZipFile:
        pass

    if not excel_okay:
        with conditions_table_p.open("r") as f:
            r = csv_tsv_reader(f, use_dict_reader=True)
            for row in r:
                items = row.items()
                print(
                    f"{row[design_matrix_sample_id_column]}: "
                    f"{', '.join(str(x[1]) for x in items if x[0] != design_matrix_sample_id_column)}"
                )
    print()

    local_output_loc = Path("./res").resolve()
    dirs = [
        local_output_loc / x
        for x in [
            "Data",
            "Data/QC",
            "Data/Contrast",
            "Plots",
            "Plots/QC",
            "Plots/QC/Variance P-Value",
            "Plots/QC/PCA",
            "Plots/Contrast",
        ]
    ]
    for x in dirs:
        x.mkdir(exist_ok=True, parents=True)

    print("\n" * 4)
    res = subprocess.Popen(
        [
            "Rscript",
            "deseq2.r",
            conditions_table_p.resolve(),
            design_matrix_sample_id_column,
            ",".join(design_formula_explanatory),
            ",".join(design_formula_confounding),
            ",".join(design_formula_cluster),
            raw_count_table_p.resolve(),
            count_table_gene_id_column,
            ",".join([]),
            str(number_of_genes_to_plot),
            str(local_output_loc),
        ],
        cwd="./r_scripts",
        stdout=subprocess.PIPE,
    )
    assert res.stdout is not None
    for l_b in res.stdout:
        l = l_b.decode("utf-8")
        if l.startswith("__LATCH_MESSAGE_DATA"):
            space1 = l.find(" ")
            space2 = l.find(" ", space1 + 1)
            typ = l[space1 + 1 : space2]
            msg = l[space2 + 1 :]
            message(typ, json.loads(msg))
            continue
        sys.stdout.write(l)

    ret_code = res.wait()
    if ret_code != 0:
        warn(f"R script failed with return code {ret_code}")
        warning(
            {
                "title": "R script failed",
                "body": "The DESeq2 run failed",
            }
        )
        raise RuntimeError("R script failed")

    if not res.stdout.closed:
        res.stdout.close()

    print("\n")

    res_p = Path("./res")
    if len(list(res_p.iterdir())) == 0:
        error(
            {
                "title": "R script produced no outputs",
                "body": "The DESeq2 run produced no outputs",
            }
        )
        raise RuntimeError("No outputs produced")

    level_options = generate_report()

    with (res_p / "Report.deseqreport").open("wb") as f:
        data_p = res_p / "Data"
        plots_p = res_p / "Plots"
        qc_plots_p = plots_p / "QC"

        embedded_data = {
            "_dds": data_p / "dds.rds",
            "sample_corr": plots_p / "Sample Correlation.html",
            "counts_heatmap": qc_plots_p / "Counts Heatmap.html",
            "size_factor_qc": qc_plots_p / "Size Factor QC.html",
            **{
                "pca/" + x.with_suffix("").name: x
                for x in (qc_plots_p / "PCA").iterdir()
                if x.suffix == ".html"
            },
        }
        embedded_data_order = sorted([k for k in embedded_data])

        json_blob = json.dumps(
            {
                "report_name": report_name,
                "genes": sorted(list(genes)),
                "level_options": level_options,
                "embedded_data_sizes": {
                    k: v.stat().st_size for k, v in embedded_data.items()
                },
                "embedded_data_order": embedded_data_order,
            }
        )
        f.write(len(json_blob).to_bytes(4, "little", signed=False))
        f.write(json_blob.encode("utf-8"))

        for k in embedded_data_order:
            x = embedded_data[k]
            with x.open("rb") as fr:
                data = fr.read1()
                while len(data) > 0:
                    f.write(data)
                    data = fr.read1()

    return LatchDir(str(res_p.resolve()), remote_path=output_loc)


@dataclass(frozen=True)
class ForkBranchDeseq2(ForkBranch):
    _tmp_unwrap_optionals: Optional[List[str]] = None

    def __init__(
        self, display_name: str, _tmp_unwrap_optionals: List[str], *flow: FlowBase
    ):
        super().__init__(display_name, *flow)
        object.__setattr__(self, "_tmp_unwrap_optionals", _tmp_unwrap_optionals)


@dataclass
class LatchParameterDeseq2(LatchParameter):
    tmp_hack_deseq2: Optional[str] = None
    add_button_title: Optional[str] = None

    @property
    def dict(self):
        res: Dict[str, Any] = super().dict

        # todo(maximsmol): add escape hatch to LatchParameter
        if self.tmp_hack_deseq2 is not None:
            res["__metadata__"]["_tmp_hack_deseq2"] = self.tmp_hack_deseq2

        # todo(maximsmol): this should really be added to LatchParameter
        if self.add_button_title is not None:
            res["__metadata__"].setdefault("appearance", {})[
                "add_button_title"
            ] = self.add_button_title

        return res


@workflow(
    metadata=LatchMetadata(
        display_name="DESeq2 (Differential Expression)",
        author=LatchAuthor(
            name="LatchBio",
            email="dev@latch.bio",
            github="https://github.com/latch-verified",
        ),
        wiki_url="https://www.latch.wiki/bulk-rna-seq-end-to-end#b2a4c0654d47450396ce094b1c70cb58",
        documentation="https://www.latch.wiki/bulk-rna-seq-end-to-end#b2a4c0654d47450396ce094b1c70cb58",
        video_tutorial="https://www.loom.com/share/46d44143a3344860b48c2c3a5e566b63",
        no_standard_bulk_execution=True,
        parameters={
            "raw_count_table": LatchParameter(
                display_name="Table File Path", batch_table_column=True
            ),
            "raw_count_tables": LatchParameter(
                display_name="Table File Paths",
                rules=[
                    LatchRule(
                        regex=r".*\.(csv|tsv)$", message="Expected a CSV or TSV file"
                    )
                ],
            ),
            "count_table_gene_id_column": LatchParameterDeseq2(
                display_name="Gene ID Column", tmp_hack_deseq2="gene_id_column"
            ),
            "report_name": LatchParameter(
                display_name="Report Name", batch_table_column=True
            ),
            "output_location_type": LatchParameter(display_name="Output Location"),
            "output_location": LatchParameter(
                display_name="Output Path", output=True, batch_table_column=True
            ),
            "conditions_source": LatchParameter(display_name="Design Matrix"),
            "manual_conditions": LatchParameterDeseq2(add_button_title="Add Condition"),
            "conditions_table": LatchParameter(
                display_name="Design Matrix", batch_table_column=True
            ),
            "design_matrix_sample_id_column": LatchParameter(
                display_name="Sample ID Column"
            ),
            "design_formula": LatchParameter(display_name="Design Formula"),
            "number_of_genes_to_plot": LatchParameterDeseq2(
                display_name="Number of Top Genes to Plot",
                add_button_title="Number of Top Genes to Plot",
            ),
            "count_table_source": LatchParameter(),
        },
        flow=[
            Section(
                "Counts Table",
                Fork(
                    "count_table_source",
                    "",
                    single=ForkBranchDeseq2(
                        "Single Table",
                        ["raw_count_table"],
                        Text(
                            dedent(
                                """
                                Table of (pseudo-)counts where columns are samples and rows are genes
                                - One of the columns must contain gene IDs
                                - A subset or all of the remaining columns can be used as samples for analysis, depending on the design matrix
                                """
                            )
                        ),
                        Params("raw_count_table", "count_table_gene_id_column"),
                    ),
                    multiple=ForkBranch(
                        "Combine Tables",
                        Text(
                            dedent(
                                """
                                Multiple tables of (pseudo-)counts where columns are samples and rows are genes
                                - One of the columns must contain gene IDs
                                - A subset or all of the remaining columns can be used as samples for analysis, depending on the design matrix
                                - Tables will be merged row-wise (new samples will be added for each gene)
                                - The first column of each table must be the gene identifier
                                - The order of genes in each table must be exactly the same
                                """
                            )
                        ),
                        Params("raw_count_tables"),
                    ),
                ),
            ),
            Section(
                "Sample Conditions (Control vs Treatment, etc.)",
                Fork(
                    "conditions_source",
                    "",
                    manual=ForkBranch("Manual Input", Params("manual_conditions")),
                    table=ForkBranchDeseq2(
                        "File",
                        [
                            "conditions_table",
                            "design_matrix_sample_id_column",
                            "design_formula",
                        ],
                        Text(
                            dedent(
                                """
                                Table with sample IDs and experimental conditions
                                """
                            )
                        ),
                        Params(
                            "conditions_table",
                            "design_matrix_sample_id_column",
                            "design_formula",
                        ),
                    ),
                ),
            ),
            Section(
                "Output Settings",
                Params("report_name"),
                Fork(
                    "output_location_type",
                    "",
                    default=ForkBranch(
                        "Default",
                        Text(
                            dedent(
                                """
                                In the data view, under
                                `/DeSeq2 (Differential Expression)/{Report Name}`
                                """
                            )
                        ),
                    ),
                    custom=ForkBranchDeseq2(
                        "Custom", ["output_location"], Params("output_location")
                    ),
                ),
            ),
        ],
    )
)
def deseq2_wf(
    report_name: str,
    count_table_source: str = "single",
    raw_count_table: Optional[
        Annotated[
            LatchFile,
            FlyteAnnotation(
                {
                    "_tmp_hack_deseq2": "counts_table",
                    "rules": [
                        {
                            "regex": r".*\.(csv|tsv|xlsx)$",
                            "message": "Expected a CSV, TSV, or XLSX file",
                        }
                    ],
                }
            ),
        ]
    ] = None,
    raw_count_tables: List[LatchFile] = [],
    count_table_gene_id_column: str = "gene_id",
    output_location_type: str = "default",
    output_location: Optional[LatchOutputDir] = None,
    conditions_source: str = "manual",
    manual_conditions: Annotated[
        List[List[str]],
        FlyteAnnotation({"_tmp_hack_deseq2": "manual_design_matrix"}),
    ] = [],
    conditions_table: Optional[
        Annotated[
            LatchFile,
            FlyteAnnotation(
                {
                    "_tmp_hack_deseq2": "design_matrix",
                    "rules": [
                        {
                            "regex": r".*\.(csv|tsv|xlsx)$",
                            "message": "Expected a CSV, TSV, or XLSX file",
                        }
                    ],
                }
            ),
        ]
    ] = None,
    design_matrix_sample_id_column: Optional[
        Annotated[str, FlyteAnnotation({"_tmp_hack_deseq2": "design_id_column"})]
    ] = None,
    design_formula: Annotated[
        List[List[str]],
        FlyteAnnotation(
            {
                "_tmp_hack_deseq2": "design_formula",
                "_tmp_hack_deseq2_allow_clustering": True,
            }
        ),
    ] = [],
    number_of_genes_to_plot: int = 30,
) -> LatchDir:
    r"""Estimate variance-mean dependence in count data from high-throughput sequencing assays and test for differential expression based on a model using the negative binomial distribution.

    Using RNA-seq to generate matrices of transcript and gene abundances has become
    a staple technique for measuring cell state.[^1] Often it is desirable to use
    statistical techniques to compare these count matrices across different
    experimental conditions to reveal genes that change.[^2]

    A software benchmark conducted by Costa Silva et. al revealed
    [DESeq2](https://bioconductor.org/packages/release/bioc/html/DESeq2.html) to be
    the most performant. For each of a list of tools, reported significant genes were
    compared against ground-truth genes derived from qRT-PCR, and DESeq2
    consistently showed the highest
    [sensitivity](https://en.wikipedia.org/wiki/Sensitivity_and_specificity) (TPR) and [accuracy](https://en.wikipedia.org/wiki/Accuracy_and_precision).

    ![table](https://user-images.githubusercontent.com/31255434/182885594-e5986335-0f3a-484d-969a-306b02aa9d82.png)


    [^1]: Stark, Rory; Grzelak, Marta; Hadfield, James (2019). RNA sequencing: the teenage years. Nature Reviews Genetics, (), –. doi:10.1038/s41576-019-0150-2
    [^2]: Costa-Silva J, Domingues D, Lopes FM (2017) RNA-Seq differential expression analysis: An extended review and a software tool. PLoS ONE 12(12): e0190152. https://doi.org/10.1371/journal.pone.0190152
    """

    return deseq2(
        count_table_source=count_table_source,
        raw_count_table=raw_count_table,
        raw_count_tables=raw_count_tables,
        count_table_gene_id_column=count_table_gene_id_column,
        report_name=report_name,
        output_location_type=output_location_type,
        output_location=output_location,
        conditions_source=conditions_source,
        manual_conditions=manual_conditions,
        conditions_table=conditions_table,
        design_matrix_sample_id_column=design_matrix_sample_id_column,
        design_formula=design_formula,
        number_of_genes_to_plot=number_of_genes_to_plot,
    )


LaunchPlan(
    deseq2_wf,
    "(Foote, 2019) Human Fibroblasts",
    dict(
        raw_count_table=LatchFile(
            "s3://latch-public/welcome/deseq2/galaxy/galaxy_counts.tsv"
        ),
        raw_count_tables=[],
        report_name="(Foote, 2019) Human Fibroblasts DESeq2 Report",
        conditions_source="table",
        conditions_table=LatchFile(
            "s3://latch-public/welcome/deseq2/galaxy/galaxy_design.csv"
        ),
        design_matrix_sample_id_column="Sample",
        design_formula=[["Condition", "explanatory"]],
    ),
)
LaunchPlan(
    deseq2_wf,
    "(Knyazev, 2021) Inflammatory Bowel Diseases",
    dict(
        raw_count_table=LatchFile(
            "s3://latch-public/welcome/deseq2/ibd/ibd_counts.csv"
        ),
        raw_count_tables=[],
        report_name="(Knyazev, 2021) Inflammatory Bowel Diseases DESeq2 Report",
        conditions_source="table",
        conditions_table=LatchFile(
            "s3://latch-public/welcome/deseq2/ibd/ibd_design.csv"
        ),
        design_matrix_sample_id_column="Sample",
        design_formula=[["Condition", "explanatory"]],
    ),
)
